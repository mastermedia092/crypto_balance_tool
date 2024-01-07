import os
from loguru import logger
from openpyxl import load_workbook, Workbook

from network import Network


class ExcelExporter:
    def __init__(self, file_path):
        self.file_path = file_path

    def export_data(self, wallet, address, chain_id):
        try:
            if os.path.exists(self.file_path):
                workbook = load_workbook(self.file_path)  # Load the existing workbook
            else:
                workbook = Workbook()  # Create a new workbook

            worksheet = self.get_or_create_worksheet(workbook, address, chain_id)

            if worksheet.max_row == 1:
                headers = [
                    "Balance",
                    "Total Supply",
                    "Symbol",
                    "Contract Address",
                    "Name",
                ]
                self.write_headers(worksheet, headers)

            self.write_asset_data(worksheet, wallet.assets)

            workbook.save(self.file_path)  # Save the workbook to the file
        except Exception as e:
            logger.error(f"Error importing data from Excel: {str(e)}")
            return None

    def write_headers(self, worksheet, headers):
        for col_num, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col_num, value=header)

    def write_asset_data(self, worksheet, assets):
        start_row = worksheet.max_row + 1
        for row_num, asset in enumerate(assets, start=start_row):
            worksheet.cell(row=row_num, column=1, value=str(asset.balance))
            worksheet.cell(row=row_num, column=2, value=str(asset.total_supply))
            worksheet.cell(row=row_num, column=3, value=asset.symbol)
            worksheet.cell(row=row_num, column=4, value=asset.contract_address)
            worksheet.cell(row=row_num, column=5, value=asset.name)

    def get_or_create_worksheet(self, workbook, address, chain_id):
        worksheet_title = f"{address}_{Network.networks.get(chain_id)}"
        if worksheet_title in workbook.sheetnames:
            worksheet = workbook[worksheet_title]  # Use the existing worksheet
        else:
            worksheet = workbook.create_sheet(
                title=worksheet_title
            )  # Create a new worksheet
        return worksheet
